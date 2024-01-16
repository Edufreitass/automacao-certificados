import random
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta

# Configuração da Faker para usar locale PT-BR
fake = Faker('pt_BR')

# Cabeçalhos
headers = [
    "Nome do Curso",
    "Nome Participante",
    "Tipo de Participação",
    "Data de Início",
    "Data de Término",
    "Carga Horária (horas)",
    "Data de Emissão do Certificado"
]

# Dados
dados = []

for _ in range(10):
    nome_participante = fake.name()
    tipo_participacao = random.choice(["Palestrante", "Organizador", "Ouvinte"])
    data_inicio = (datetime.now() - timedelta(days=random.randint(1, 365))).strftime("%d/%m/%Y")
    data_termino = (datetime.strptime(data_inicio, "%d/%m/%Y") + timedelta(days=random.randint(1, 30))).strftime("%d/%m/%Y")
    carga_horaria = random.randint(1, 40)
    data_emissao_certificado = (datetime.strptime(data_termino, "%d/%m/%Y") + timedelta(days=random.randint(1, 60))).strftime("%d/%m/%Y")

    dados.append([
        "Introdução a Programação",
        nome_participante,
        tipo_participacao,
        data_inicio,
        data_termino,
        carga_horaria,
        data_emissao_certificado
    ])

# Criar planilha
wb = Workbook()
ws = wb.active

# Escrever cabeçalhos em negrito
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)

# Preencher dados
for row_num, row_data in enumerate(dados, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Salvar planilha
wb.save("planilha_alunos.xlsx")
print("Planilha criada com sucesso!")
